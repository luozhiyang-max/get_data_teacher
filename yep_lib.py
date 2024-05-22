import requests
from bs4 import BeautifulSoup
import openpyxl as op
import re
import dashscope  #调用qwenAPI
dashscope.api_key='sk-0bbdc339fabb4b5bbc6e6f9337ca7991'


#函数名称 Req_url
#输出参数：网页的url
#输出：若成功，则返回网页解析内容
#     若失败，则返回字符串“failed”
def Req_url(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0"
        }
    response = requests.get(url, headers=headers)
    response.encoding = "utf-8"  # 防止乱码
    if response.ok:
        content = response.text
        soup = BeautifulSoup(content, "html.parser")
        print("请求成功")
        return soup
    else:
        print("请求失败,状态码为 " + str(response.status_code))
        return "failed"


#输入：网页解析内容
#输出：该网页中所有的文字
def All_text(soup):
    texts=soup.find_all(text=True)
    cleaned_text = ''
    for text in texts:
        cleaned_text += re.sub(r'[A-Za-z]+', '', text)        #去除英文
        cleaned_text = re.sub(r'\s{2,}', ' ', cleaned_text)
        cleaned_text = re.sub(r'\n{2,}', '\n', cleaned_text)
    return cleaned_text


#这个函数每调用一次就在已有excel文件基础上填写一行内容
#excel_name 为已创建的表格名称，要求与.py文件在同一路径下
#judge_text:是qwen生成的文本信息
#bias为列的偏移量，按照蔡哥的表格来看偏移量为2，前两列为地区和大学
def W_excel(excel_name,judge_text,bias):
    workbook = op.load_workbook(excel_name)
    # 选择要操作的工作表
    sheet = workbook.active
    # 找到下一个空白行
    next_row = sheet.max_row + 1
    # 在下一个空白行写入信息
    lines = judge_text.split('\n')
    sheet.cell(row=next_row, column=bias+1, value=lines[0][3:])
    sheet.cell(row=next_row, column=bias+2, value=lines[1][3:])
    sheet.cell(row=next_row, column=bias+3, value=lines[2][3:])
    sheet.cell(row=next_row, column=bias+4, value=lines[3][3:])
    sheet.cell(row=next_row, column=bias+5, value=lines[4][3:])
    sheet.cell(row=next_row, column=bias+6, value=lines[5][3:])
    sheet.cell(row=next_row, column=bias+7, value=lines[6][3:])
    sheet.cell(row=next_row, column=bias+8, value=lines[7][3:])
    sheet.cell(row=next_row, column=bias+10, value=lines[8][3:])
    sheet.cell(row=next_row, column=bias+11, value=lines[9][3:])
    sheet.cell(row=next_row, column=bias+12, value=lines[10][4:])
    sheet.cell(row=next_row, column=bias+13, value=lines[11][4:])
    sheet.cell(row=next_row, column=bias+15, value=lines[12][4:])
    sheet.cell(row=next_row, column=bias+16, value=lines[13][4:])
    sheet.cell(row=next_row, column=bias+17, value=lines[14][4:])
    sheet.cell(row=next_row, column=bias+18, value=lines[15][4:])
    sheet.cell(row=next_row, column=bias+19, value=lines[16][4:])
    sheet.cell(row=next_row, column=bias+20, value=lines[17][4:])
    sheet.cell(row=next_row, column=bias+21, value=lines[18][4:])
    sheet.cell(row=next_row, column=bias+22, value=lines[19][4:])
    # 保存 Excel 文件
    workbook.save(excel_name)


#qwen_max API
def Qwen_model(texts):
    judge = dashscope.Generation.call(
        model=dashscope.Generation.Models.qwen_max,
        prompt=f'''以下文本片段是从网页中爬取的一个网页中全部的文本信息，注意可能会出现无关的文本信息，请分辨出其中与教师信息相关的文本，并根据给出的标准完成以下n个任务：

        文本片段：{texts}
        1. 输出该文本中对应教师的名字
        2. 输出该文本中任务的职称，例如院长，副院长，主任，教授，副教授，研究员等
        3. 是否有职称，如果有输出1，没有输出0
        4. 输出文本中提到的荣誉称号，例如**市**人才，四青(优青、海优、青长、青拔)+长江/万人/杰青/领军，海外高层次，国家高层次，科技部领军、国家百千万人才工程...奖励:国家自然科学/技术发明/科学进步奖、科技部/交通部等等省市级/学会级/国家级奖励等
        5. 是否为青年长江学者，是的话输出1，没提及输出0
        6. 是否为优秀青年基金，是的话输出1，没提及输出0
        7. 是否为青年拔尖人才，是的话输出1，没提及输出0
        8. 是否为青年千人（又称海外高层次人才引进计划），是的话输出1，没提及输出0
        9. 上述第4，第5，第6，第7条标准是否有一条满足，是的话输出1，不是的话输出0
        10. 是否为杰青（即国家杰出青年科学基金），是的话输出1，没提及输出0
        11. 是否为长江学者（即长江学者奖励计划），注意与第4条青年长江学者不是同一类型，是的话输出1，没提及输出0
        12. 是否为万人计划，是的话输出1，没提及输出0
        13. 上述第9，第10，第11条标准是否有一条满足，是的话输出1，不是的话输出0
        14. 是否为科学院院士，是的话输出1，不是的话输出0
        15. 是否为工程院院士，是的话输出1，不是的话输出0
        16. 是否为其他院院士，是的话输出具体信息，不是的话输出0
        17. 是否为IEEE Fellow，是的话输出1，不是的话输出0
        18. 是否为其他Fellow, 是的话输出具体Fellow信息，不是的话输出0
        19. 输出其主要研究方向，没提及的话输出信息中没提及
        20. 输出其曾获项目，包括时间段以及项目名称，没提及的话输出没提及
        请注意，如果内容无需修改，请完整呈现文本片段不要省略任何内容，如果文本对话内容需要修改，请整合出修改后的完整对话，输出的内容不需要掺杂额外内容例如审核结果，内容分析等，只需给出原本格式的对话格式即可。
        ''')
    judge_dialog = judge["output"]["text"]
    return judge_dialog


